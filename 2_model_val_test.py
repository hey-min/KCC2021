# -*- coding: utf-8 -*-
# 2021.06.02

import os
import sys
import time
import logging
import argparse

import IPython
import IPython.display
import numpy as np
import pandas as pd
import tensorflow as tf
import pickle
import math
from openpyxl import Workbook, load_workbook

tf.compat.v1.reset_default_graph()

parser = argparse.ArgumentParser()
parser.add_argument('--EST', type=int, default=1, dest='EST')
parser.add_argument('--LR', type=float, default=0.0005, dest='LR')
parser.add_argument('--IT', type=int, default=500, dest='IT')
# parser.add_argument('--BATCH', type=int, default=14, dest='BATCH')
args = parser.parse_args()


EST = args.EST
LR = args.LR
IT = args.IT
# BATCH = args.BATCH
VER = str(EST)+'_'+str(LR)+'_'+str(IT)



val_performance = {}
performance = {}


def createFolder(dir):
    try:
        if not os.path.exists(dir):
            os.makedirs(dir)
    except Exception as e:
        print(e)
        
    

NAME_PICKLE = '2007to2019.pickle'
def read_pickle():
    
    # get pickle
    with open(NAME_PICKLE, 'rb') as f:
        nc_sst = pickle.load(f)
        nc_lat = pickle.load(f)
        nc_lon = pickle.load(f)
        nc_time = pickle.load(f)
    
    return nc_sst, nc_lat, nc_lon, nc_time

def get_data(lat, lon):
    
    total_size = len(nc_time)
    columns = ['sst']
    df = pd.DataFrame(columns=columns)
    
    for i in range(total_size):
        new_sst = np.round(nc_sst[i, lat, lon] - 273.15, 4)
        # new_date= str(nc_time[i])[:10]
        df_data = {'sst':new_sst}
        df = df.append(pd.Series(df_data, index=df.columns), ignore_index=True)
    
    return df





######################################################################
####################  CLASS :  WindowGenerator   #####################
######################################################################


class WindowGenerator():
    
    def __init__(self, input_width, label_width, shift,
               train_df, val_df, test_df,
               label_columns=None):
        # Store the raw data.
        self.train_df = train_df
        self.val_df = val_df
        self.test_df = test_df
        
        # Work out the label column indices.
        self.label_columns = label_columns
        if label_columns is not None:
            self.label_columns_indices = {name: i for i, name in
                                    enumerate(label_columns)}
        self.column_indices = {name: i for i, name in
                           enumerate(train_df.columns)}

        # Work out the window parameters.
        self.input_width = input_width
        self.label_width = label_width
        self.shift = shift
        
        self.total_window_size = input_width + shift
        
        self.input_slice = slice(0, input_width)
        self.input_indices = np.arange(self.total_window_size)[self.input_slice]
        
        self.label_start = self.total_window_size - self.label_width
        self.labels_slice = slice(self.label_start, None)
        self.label_indices = np.arange(self.total_window_size)[self.labels_slice]

    def __repr__(self):
        return '\n'.join([
            f'Total window size: {self.total_window_size}',
            f'Input indices: {self.input_indices}',
            f'Label indices: {self.label_indices}',
            f'Label column name(s): {self.label_columns}'])

    def split_window(self, features):

        inputs = features[:, self.input_slice, :]
        labels = features[:, self.labels_slice, :]
        

        if self.label_columns is not None:
            labels = tf.stack(
                [labels[:, :, self.column_indices[name]] for name in self.label_columns], axis=-1)

        inputs.set_shape([None, self.input_width, None])
        labels.set_shape([None, self.label_width, None])

        return inputs, labels

   
    
    def plot(self, model=None, plot_col='sst', max_subplots=1):
        
        inputs, labels = self.example
        plot_col_index = self.column_indices[plot_col]
        max_n = min(max_subplots, len(inputs))
        
        for n in range(max_n):
            
            if self.label_columns:
                label_col_index = self.label_columns_indices.get(plot_col, None)
            else:
                label_col_index = plot_col_index

            if label_col_index is None:
                continue

            if model is not None:
                predictions = model(inputs)


        data_input = inputs.numpy()
        data_output = predictions.numpy()
        
        return data_input, data_output


    def make_dataset(self, data):
        
        data = np.array(data, dtype=np.float32)
        ds = tf.keras.preprocessing.timeseries_dataset_from_array(
            data=data,
            targets=None,
            sequence_length=self.total_window_size,
            sequence_stride=1,
            shuffle=True,
            batch_size=30,)

        ds = ds.map(self.split_window)
      
        return ds


@property
def train(self):
  # print('TRAIN START')
  return self.make_dataset(self.train_df)

@property
def val(self):
  # print('VALIDATION START')
  return self.make_dataset(self.val_df)

@property
def test(self):
  # print('TEST START')
  return self.make_dataset(self.test_df)

@property
def example(self):
  """Get and cache an example batch of `inputs, labels` for plotting."""
  result = getattr(self, '_example', None)
  if result is None:
    # No example batch was found, so get one from the `.train` dataset
    result = next(iter(self.test))
    # And cache it for next time
    self._example = result
  return result

WindowGenerator.train = train
WindowGenerator.val = val
WindowGenerator.test = test
WindowGenerator.example = example



MAX_EPOCHS = IT
def compile_and_fit(model, window, patience=2):
  early_stopping = tf.keras.callbacks.EarlyStopping(monitor='val_loss',
                                                    patience=patience,
                                                    mode='min')

  model.compile(loss=tf.losses.MeanSquaredError(),
                optimizer=tf.optimizers.Adam(lr=LR),
                metrics=[tf.metrics.MeanAbsoluteError()])

  history = model.fit(window.train, epochs=MAX_EPOCHS,
                      validation_data=window.val,
                      verbose=2,
                      callbacks=[early_stopping])
  return history







def ValTest(r, c):
    
    LAT = r
    LON = c
    
    model_path = 'model/'+VER
    # createFolder(model_path)
    model_name = '['+str(LAT)+']['+str(LON)+']'
    
    print('est:{} lat:{} lon:{} lr:{} it:{}' .format(EST, LAT, LON, LR, IT))
  

    df = get_data(LAT, LON)

    column_indices = {name: i for i, name in enumerate(df.columns)}

    n = len(df)
    train_df = df[0:3653]
    val_df = df[3653:4383]

    find_date = '2019-08-10 12:00:00'
    df_date = pd.DataFrame(nc_time, columns={'date'})
    find_index = df_date[df_date['date']==find_date].index.values
    
    
    LABEL = int(find_index)
    END = int(LABEL - EST)
    START = int(END - 30 + 1)
    test_df = df[START:LABEL+1]
    
    ############# CONV WINDOW #############
    CONV_WIDTH = 30
    conv_window = WindowGenerator(
        input_width=CONV_WIDTH,
        label_width=1,
        shift=EST,
        label_columns=['sst'],
        train_df=train_df, val_df=val_df, test_df=test_df)


    ############# CONV MODEL #############
    conv_model = tf.keras.Sequential([
        tf.keras.layers.Conv1D(filters=32,
                               kernel_size=(CONV_WIDTH,),
                               activation='relu'),
        tf.keras.layers.Dense(units=32, activation='relu'),
        tf.keras.layers.Dense(units=1),
    ])



    ############# WIDE WINDOW #############
    wide_window = WindowGenerator(
        input_width=30, label_width=30, shift=EST,
        label_columns=['sst'],
        train_df=train_df, val_df=val_df, test_df=test_df)




    ############# LSTM MODEL #############
    lstm_model = tf.keras.models.Sequential([
        tf.keras.layers.LSTM(32, return_sequences=True),
        # Shape => [batch, time, features]
        tf.keras.layers.Dense(units=1)
    ])
    
    print('MODEL:{}' .format(model_path+'/'+model_name+'.h5'))
    new_model = tf.keras.models.load_model(model_path+'/'+model_name+'.h5')
    
    # print('\n-----------------MODEL VALIDATION-----------------')
    ds_val = wide_window.val
    val_performance['LSTM'] = new_model.evaluate(ds_val, verbose=2)
    # performance : loss, mse
    list_val = list(ds_val.as_numpy_iterator())

    
    # print('\n-------------------MODEL TEST-------------------')
    ds_test = wide_window.test
    performance['LSTM'] = new_model.evaluate(ds_test, verbose=2)

    list_test = list(ds_test.as_numpy_iterator())

    rslt_input = list_test[0][0]
    rslt_label = list_test[0][1]



    rslt_input, rslt_output = wide_window.plot(new_model)
    rslt_df = pd.DataFrame({'index':test_df.index[:30],
                            'input': rslt_input.reshape(30),
                            'label' : rslt_label.reshape(30),
                            'predict':rslt_output.reshape(30)})

    
    est_sst = rslt_output[0][29]
    
    
    
    return est_sst

if __name__ == '__main__':
    
    nc_sst, nc_lat, nc_lon, nc_time = read_pickle()
    
    excel = 'excel/EST'+VER+'.xlsx'
    
    version = 0
    while os.path.isfile(excel):
        version = version + 1
        excel = 'excel/EST'+VER+'_v'+str(version)+'.xlsx'
    
    wb = Workbook()
    ws = wb.active
         
    for r in range(len(nc_lat)):
        for c in range(len(nc_lon)):
        
            data = nc_sst[0][r][c]
            
            if math.isnan(data):
                continue
            
            else:
                est = ValTest(r, c)
                rslt = round(est[0], 3)
                print('Est: {}' .format(rslt))
                
                ws.cell(r+1, c+1, rslt)
    
    ws.insert_rows(1)
    wb.save(excel)
    wb.close()




















